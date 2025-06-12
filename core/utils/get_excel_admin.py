import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from middlewares.enums import Variables


def export_all_user_results_to_excel(variables: Variables) -> BytesIO:
    test_result_repo = variables.db.test_result
    user_repo = variables.db.user

    results = test_result_repo.get_all_with_users()  # 👈 ты должен реализовать эту функцию в репозитории

    data = []
    for result, user in results:
        data.append({
            "ID": str(user.user_id),
            "ФИО": user.full_name,
            "Email": user.email or '',
            "Username": user.username or '',
            "Организация": user.educational_organization or '',
            "Статус": user.status or '',
            "Конгресс": "Да" if user.congress else "Нет",
            "Тест": result.test_name,
            "Правильных": result.correct_answers,
            "Всего": result.total_questions,
            "Баллы": result.points
        })

    df = pd.DataFrame(data)

    output = BytesIO()
    df.to_excel(output, index=False, sheet_name="Результаты")
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center_align

    ws.column_dimensions['A'].width = 14.3
    ws.column_dimensions['B'].width = 28.5
    ws.column_dimensions['C'].width = 28.5
    ws.column_dimensions['D'].width = 14.3
    ws.column_dimensions['E'].width = 14.3
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 14.3
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 14.3
    ws.column_dimensions['K'].width = 14.3

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "@"

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output
